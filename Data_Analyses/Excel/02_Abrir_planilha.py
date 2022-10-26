
#! ABRIR E MODIFICAR EXCEL

import openpyxl

book = openpyxl.load_workbook(r'C:\Users\goncavic01\.vscode\github\Python\Python-1\SDA_Vendas_python.xlsx')

vendas_page = book['Vendas']

#* iterando sobre todas as linhas
# start row 2 and finish row 5
def print_basico():
    for rows in vendas_page.iter_rows(min_row=2):
        for cell in rows:
            print(cell.value)           #somente cell mostra o caminho A4,B4...

#* print na mesma linha
def print_linha():
    for rows in vendas_page.iter_rows(min_row=2):
        print()
        print(f'{rows[0].value};{rows[1].value};{rows[2].value}')

#* MODIFICAR UMA PLANILHA
for rows in vendas_page.iter_rows(min_row=2):
    for cell in rows:
        if cell.value == 'Via':
            cell.value = 'VIA VAREJO'

#* print na mesma linha
print_linha()

book.save('SDA_Vendas_python.xlsx')