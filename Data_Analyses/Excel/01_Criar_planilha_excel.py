
#! CRIR UMA PLANILHA EXCEL DO ZERO


import openpyxl

#Criar uma planilha
book = openpyxl.Workbook()

#Visualizar as paginas do book
# print(book.sheetnames)

#Criar uma pagina
book.create_sheet('Vendas')
page_vendas = book['Vendas']

#adicionar dados

page_vendas.append(['Cliente','Modelo','Sell Out',"Estoque"])
page_vendas.append(['Via','STK12',1000,20000])
page_vendas.append(['Luiza','STK12',800,17000])
page_vendas.append(['Havan','STK12',500,15000])

#Salvar a planilha
book.save('SDA_Vendas_python.xlsx')